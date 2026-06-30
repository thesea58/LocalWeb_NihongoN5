"""Import vocabulary lessons 10-25 from the project workbook into JSON files."""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_ROOT / "tu vung bai 10-25.xlsx"
OUTPUT_DIRECTORY = PROJECT_ROOT / "data" / "lessons"
SHEET_NAME = "10-25"
EXPECTED_HEADERS = [
    "Bài",
    "Từ tiếng Nhật",
    "Kanji",
    "Loại từ",
    "Nghĩa tiếng Việt",
]
LESSONS = range(10, 26)


def clean_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def read_lessons() -> dict[int, list[dict[str, object]]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Không tìm thấy sheet {SHEET_NAME!r}.")

    worksheet = workbook[SHEET_NAME]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Tiêu đề cột không hợp lệ: {headers!r}")

    lessons: dict[int, list[dict[str, object]]] = defaultdict(list)
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not any(value is not None for value in values):
            continue

        lesson_value, japanese_value, kanji_value, word_type_value, meaning_value = (
            values
        )
        lesson = int(lesson_value)
        japanese = clean_text(japanese_value)
        kanji = clean_text(kanji_value)
        word_type = clean_text(word_type_value)
        meaning = clean_text(meaning_value)

        if lesson not in LESSONS:
            raise ValueError(f"Dòng {row_number}: bài {lesson} nằm ngoài phạm vi 10-25.")
        if not japanese or not word_type or not meaning:
            raise ValueError(f"Dòng {row_number}: thiếu dữ liệu bắt buộc.")

        lessons[lesson].append(
            {
                "lesson": lesson,
                "kanji": kanji or japanese,
                "hira": japanese,
                "wordType": word_type,
                "meaning": meaning,
            }
        )

    missing_lessons = [lesson for lesson in LESSONS if not lessons[lesson]]
    if missing_lessons:
        raise ValueError(f"Không có dữ liệu cho các bài: {missing_lessons}")

    return dict(lessons)


def write_lessons(lessons: dict[int, list[dict[str, object]]]) -> None:
    OUTPUT_DIRECTORY.mkdir(parents=True, exist_ok=True)
    for lesson in LESSONS:
        output_path = OUTPUT_DIRECTORY / f"lesson-{lesson}.json"
        content = json.dumps(lessons[lesson], ensure_ascii=False, indent=2) + "\n"
        output_path.write_text(content, encoding="utf-8")
        print(f"{output_path.relative_to(PROJECT_ROOT)}: {len(lessons[lesson])} records")


if __name__ == "__main__":
    write_lessons(read_lessons())
