"""Import JLPT N5 kanji workbook into app vocabulary JSON."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_ROOT / "Han_tu_N5_JLPT_2010_2025.xlsx"
OUTPUT_PATHS = [
    PROJECT_ROOT / "data" / "kanji-n5-jlpt-2010-2025.json",
    PROJECT_ROOT.parent
    / "webapp-hoc-tieng-nhat"
    / "FlashCard"
    / "data"
    / "kanji-n5-jlpt-2010-2025.json",
]
EXPECTED_HEADERS = [
    "Kỳ thi",
    "Phần",
    "Hán tự",
    "Cách đọc",
    "Hán Việt",
    "Ý nghĩa",
    "Trang ảnh",
]


def clean_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def read_kanji() -> list[dict[str, object]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)[: len(EXPECTED_HEADERS)]]
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Tiêu đề cột không hợp lệ: {headers!r}")

    words: list[dict[str, object]] = []
    for row_number, values in enumerate(rows, start=2):
        values = list(values[: len(EXPECTED_HEADERS)])
        if not any(value is not None for value in values):
            continue

        exam, section, kanji, reading, sino_vietnamese, meaning, image_page = values
        entry = {
            "kanji": clean_text(kanji),
            "hira": clean_text(reading),
            "wordType": clean_text(sino_vietnamese),
            "meaning": clean_text(meaning),
            "exam": clean_text(exam),
            "section": clean_text(section),
            "imagePage": image_page,
        }

        missing_fields = [
            key for key in ("kanji", "hira", "meaning", "exam", "section") if not entry[key]
        ]
        if missing_fields:
            raise ValueError(f"Dòng {row_number}: thiếu {', '.join(missing_fields)}.")

        words.append(entry)

    if not words:
        raise ValueError("Workbook không có dữ liệu Kanji.")

    return words


def write_kanji(words: list[dict[str, object]]) -> None:
    content = json.dumps(words, ensure_ascii=False, indent=2) + "\n"
    for output_path in OUTPUT_PATHS:
        if not output_path.parent.exists():
            continue
        output_path.write_text(content, encoding="utf-8")
        print(f"{output_path.relative_to(PROJECT_ROOT.parent)}: {len(words)} records")


if __name__ == "__main__":
    write_kanji(read_kanji())
